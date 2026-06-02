import openpyxl

def inspect_data_colors(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    print("--- Row 11 Cell Colors ---")
    for col in range(1, 18):
        col_letter = openpyxl.utils.get_column_letter(col)
        cell = sheet.cell(row=11, column=col)
        fill = cell.fill
        color_hex = None
        if fill and isinstance(fill, openpyxl.styles.PatternFill) and fill.fill_type == 'solid':
            if fill.fgColor and fill.fgColor.rgb:
                color_hex = fill.fgColor.rgb
        print(f"Col {col_letter} ({sheet.cell(10, col).value}): Color={color_hex}")

if __name__ == "__main__":
    inspect_data_colors(r"c:\Portifólio\pi-web\Docs\New\PROFORMA- SADIMA FRG2026-PO-04.xlsx")
