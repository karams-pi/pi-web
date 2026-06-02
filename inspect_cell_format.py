import openpyxl

def inspect_format(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb.active
    
    print("--- Row 10 Cell Styling details ---")
    for col in range(1, 18):
        col_letter = openpyxl.utils.get_column_letter(col)
        cell = sheet.cell(row=10, column=col)
        
        fill = cell.fill
        fill_type = fill.fill_type if fill else None
        fg_color = fill.fgColor.rgb if fill and fill.fgColor else None
        bg_color = fill.bgColor.rgb if fill and fill.bgColor else None
        theme = fill.fgColor.theme if fill and fill.fgColor else None
        tint = fill.fgColor.tint if fill and fill.fgColor else None
        
        font = cell.font
        font_name = font.name if font else None
        font_size = font.size if font else None
        font_color = font.color.rgb if font and font.color else None
        font_bold = font.bold if font else None
        
        print(f"Col {col_letter}: Header='{cell.value}'")
        print(f"  Fill: type={fill_type}, fgColor={fg_color}, bgColor={bg_color}, theme={theme}, tint={tint}")
        print(f"  Font: name={font_name}, size={font_size}, color={font_color}, bold={font_bold}")

if __name__ == "__main__":
    inspect_format(r"c:\Portifólio\pi-web\Docs\New\PROFORMA- SADIMA FRG2026-PO-04.xlsx")
