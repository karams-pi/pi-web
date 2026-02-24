import openpyxl
import re

wb = openpyxl.load_workbook('Docs/TESTE1.xlsx', data_only=True)
ws = wb.active # Using active sheet

print(f"Sheet: {ws.title}")
for row_idx in range(1, 15): # Inspect first 15 rows
    row_data = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 16)]
    print(f"Row {row_idx}: {row_data}")

# Specific check for DAMMAN
found = False
for row_idx in range(1, ws.max_row + 1):
    val = ws.cell(row=row_idx, column=2).value
    if val and 'DAMMAN' in str(val).upper():
        found = True
        row_data = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 16)]
        print(f"Found DAMMAN at Row {row_idx}: {row_data}")
        # Look at next few rows
        for i in range(1, 5):
            next_row = [ws.cell(row=row_idx+i, column=col_idx).value for col_idx in range(1, 16)]
            print(f"  Next Row {row_idx+i}: {next_row}")
        break

if not found:
    print("DAMMAN not found in first column. Checking column 12 (Description)...")
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=12).value
        if val and 'DAMMAN' in str(val).upper():
            row_data = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 16)]
            print(f"Found DAMMAN (Desc) at Row {row_idx}: {row_data}")
            break
