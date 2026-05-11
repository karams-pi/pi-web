import pandas as pd
import openpyxl

file_path = r'c:\Portifólio\seawyse\EDC\Docs\EDC - RALLI - 11-03-2026.xlsx'
output_path = r'c:\Portifólio\seawyse\EDC\Docs\edc_analysis_dump.txt'

def analyze_excel(file_path, output_path):
    try:
        # Load the workbook to get sheet names and potentially formulas
        wb = openpyxl.load_workbook(file_path, data_only=False)
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(f"ANALYSIS OF EDC FILE: {file_path}\n")
            f.write("="*50 + "\n\n")
            
            f.write(f"Sheets found: {wb.sheetnames}\n\n")
            
            for sheet_name in wb.sheetnames:
                f.write(f"--- SHEET: {sheet_name} ---\n")
                sheet = wb[sheet_name]
                sheet_data = wb_data[sheet_name]
                
                # Get dimensions
                max_row = sheet.max_row
                max_col = sheet.max_column
                f.write(f"Dimensions: {max_row} rows x {max_col} columns\n\n")
                
                # Sample data (first 100 rows)
                f.write("Data Sample (First 100 rows):\n")
                for r in range(1, min(max_row + 1, 101)):
                    row_vals = []
                    for c in range(1, max_col + 1):
                        val = sheet_data.cell(row=r, column=c).value
                        row_vals.append(str(val) if val is not None else "")
                    f.write("\t".join(row_vals) + "\n")
                
                f.write("\nFormulas found in this sheet:\n")
                for r in range(1, min(max_row + 1, 200)): # Check first 200 rows for formulas
                    for c in range(1, max_col + 1):
                        cell = sheet.cell(row=r, column=c)
                        if cell.data_type == 'f':
                            f.write(f"Cell {cell.coordinate}: {cell.value}\n")
                
                f.write("\n" + "-"*50 + "\n\n")
                
        print(f"Success! Analysis dumped to {output_path}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    analyze_excel(file_path, output_path)
