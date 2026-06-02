import openpyxl

def inspect_excel(file_path, output_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f"Sheet Name: {sheet.title}\n")
        f.write(f"Max Row: {sheet.max_row}, Max Column: {sheet.max_column}\n\n")
        
        # Dimensions / Column Widths
        f.write("--- Column Widths ---\n")
        for col in range(1, sheet.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            width = sheet.column_dimensions[col_letter].width
            f.write(f"Column {col_letter}: {width}\n")
        f.write("\n")
        
        # Merged cells
        f.write("--- Merged Ranges ---\n")
        for merged_range in sheet.merged_cells.ranges:
            f.write(f"{merged_range}\n")
        f.write("\n")
        
        # Cell Values and Types
        f.write("--- Cell Values and Formats ---\n")
        for r in range(1, sheet.max_row + 1):
            row_vals = []
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                val = cell.value
                val_str = ""
                if val is not None:
                    # Let's see if the cell is merged
                    val_str = f"{val}"
                    if len(val_str) > 50:
                        val_str = val_str[:47] + "..."
                    val_str = f"'{val_str}'"
                    # Add type and format info
                    val_str += f" (Format: {cell.number_format})"
                else:
                    val_str = "None"
                row_vals.append(f"{openpyxl.utils.get_column_letter(c)}{r}: {val_str}")
            
            # Print if not all None
            if any(cell.value is not None for cell in [sheet.cell(row=r, column=col) for col in range(1, sheet.max_column + 1)]):
                f.write(f"Row {r:02d}: {', '.join(row_vals)}\n")

if __name__ == "__main__":
    import sys
    inspect_excel(
        r"c:\Portifólio\pi-web\Docs\New\PROFORMA- SADIMA FRG2026-PO-04.xlsx",
        r"c:\Portifólio\pi-web\excel_dump.txt"
    )
    print("Done inspection.")
