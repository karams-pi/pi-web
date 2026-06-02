import openpyxl

def compare():
    # Load user reference
    wb_ref = openpyxl.load_workbook(r"c:\Portifólio\pi-web\Docs\New\PROFORMA- SADIMA FRG2026-PO-04.xlsx", data_only=True)
    sheet_ref = wb_ref.active
    
    # Load generated
    wb_gen = openpyxl.load_workbook(r"c:\Portifólio\pi-web\test_export_15.xlsx", data_only=True)
    sheet_gen = wb_gen.active
    
    print("Column | Reference Header | Generated Header")
    print("---------------------------------------------")
    max_cols = max(sheet_ref.max_column, sheet_gen.max_column)
    for col in range(1, max_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        ref_val = sheet_ref.cell(row=10, column=col).value
        gen_val = sheet_gen.cell(row=10, column=col).value
        print(f"Col {col_letter:2s}  | {str(ref_val):16s} | {str(gen_val)}")

if __name__ == "__main__":
    compare()
