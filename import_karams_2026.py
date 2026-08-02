import psycopg2
import openpyxl
import re
import unicodedata

def normalize(text):
    if not text:
        return ""
    nfkd_form = unicodedata.normalize('NFKD', text)
    text = "".join([c for c in nfkd_form if not unicodedata.combining(c)])
    text = text.lower().strip()
    text = re.sub(r'\s+', ' ', text)
    text = text.replace('´', "'").replace('’', "'")
    return text

def clean_val(val):
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val == "" or val.lower() == "none" or val == "-" or val == "0":
            return None
        val = val.replace(',', '.')
        try:
            return float(val)
        except ValueError:
            return val
    return float(val)

def clean_desc(desc):
    if desc is None:
        return ""
    desc = str(desc).strip()
    desc = re.sub(r'\s+', ' ', desc)
    return desc

def main():
    excel_path = r"c:\Portifólio\pi-web\Docs\New\Karams 2026 - Abimad 42 (Exportação).xlsm"
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # 1. Connect to local DB to check current state
    conn = psycopg2.connect("host=localhost port=5432 dbname=pi_db user=pi password=pi123")
    cur = conn.cursor()
    
    # Cache categories
    cur.execute("SELECT id, nome FROM pi.categoria;")
    cat_map = {normalize(r[1]): r[0] for r in cur.fetchall()}
    
    # Cache brands
    cur.execute("SELECT id, nome FROM pi.marca;")
    brand_map = {normalize(r[1]): r[0] for r in cur.fetchall()}
    
    # Cache existing modules for supplier = 1 (Karams)
    cur.execute("""
        SELECT m.id, m.id_marca, m.descricao, m.largura, m.profundidade, m.id_categoria
        FROM pi.modulo m
        WHERE m.id_fornecedor = 1;
    """)
    db_modules = []
    for r in cur.fetchall():
        db_modules.append({
            "id": r[0],
            "id_marca": r[1],
            "descricao_raw": r[2],
            "descricao_norm": normalize(r[2]),
            "largura": float(r[3]),
            "profundidade": float(r[4]),
            "id_categoria": r[5]
        })
    print(f"Loaded {len(db_modules)} existing Karams modules from DB.")
    
    # Cache fabrics
    cur.execute("SELECT id, nome FROM pi.tecido;")
    tecido_map = {normalize(r[1]): r[0] for r in cur.fetchall()}
    
    # Sheets to categories mapping
    # Note sheet names and exact DB names
    sheet_cat_mapping = {
        "Karam´s Estofados": "Karam´s Estofados",
        "Karam´s Poltronas": "Karam´s Poltronas",
        "Karam´s Puff e Almofadas": "Karam´s Puff e Almofadas",
        "Karam´s Camas": "Karam´s Camas"
    }
    
    # Brands and fabrics we will check for insertion
    brands_to_ensure = set()
    fabrics_to_ensure = set()
    
    parsed_modules = []
    
    for sheet_name, db_cat_name in sheet_cat_mapping.items():
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet {sheet_name} not found in workbook.")
            continue
        sheet = wb[sheet_name]
        
        cat_norm = normalize(db_cat_name)
        cat_id = cat_map.get(cat_norm)
        if not cat_id:
            raise Exception(f"Category '{db_cat_name}' not found in DB!")
            
        current_model = None
        col_mapping = {}
        
        for r in range(1, sheet.max_row + 1):
            b_val = sheet.cell(row=r, column=2).value
            b_clean = clean_desc(b_val)
            
            # Check if it is a header row
            if b_clean and any(x in b_clean.lower() for x in ["descrição", "descriço"]):
                a_val = sheet.cell(row=r, column=1).value
                if a_val:
                    current_model = clean_desc(a_val)
                    brands_to_ensure.add(current_model)
                    
                col_mapping = {}
                for col in range(8, 19):
                    cell_val = sheet.cell(row=r, column=col).value
                    if cell_val:
                        cell_str = str(cell_val).strip().upper()
                        if re.match(r'^G\d+$', cell_str):
                            col_mapping[col] = cell_str
                            fabrics_to_ensure.add(cell_str)
                continue
                
            if not current_model:
                continue
                
            # Check if it is a module row
            c_val = sheet.cell(row=r, column=3).value
            if b_clean != "" and c_val is not None:
                width = clean_val(c_val)
                if not isinstance(width, float):
                    continue
                    
                prof = clean_val(sheet.cell(row=r, column=4).value)
                pa = 0.0
                m3 = 0.0
                altura = 0.0
                
                if sheet_name == "Karam´s Estofados":
                    pa = clean_val(sheet.cell(row=r, column=5).value) or 0.0
                    m3 = clean_val(sheet.cell(row=r, column=6).value) or 0.0
                    altura = clean_val(sheet.cell(row=r, column=7).value) or 0.0
                elif sheet_name in ["Karam´s Poltronas", "Karam´s Puff e Almofadas"]:
                    altura = clean_val(sheet.cell(row=r, column=5).value) or 0.0
                    m3 = clean_val(sheet.cell(row=r, column=6).value) or 0.0
                elif sheet_name == "Karam´s Camas":
                    m3 = clean_val(sheet.cell(row=r, column=5).value) or 0.0
                    altura = clean_val(sheet.cell(row=r, column=6).value) or 0.0
                    
                # Read prices
                prices = {}
                for col, fab_name in col_mapping.items():
                    price = clean_val(sheet.cell(row=r, column=col).value)
                    if isinstance(price, float) and price > 0:
                        prices[fab_name] = price
                        
                parsed_modules.append({
                    "sheet": sheet_name,
                    "row": r,
                    "cat_id": cat_id,
                    "brand": current_model,
                    "description": b_clean,
                    "width": width,
                    "prof": prof,
                    "altura": altura,
                    "pa": pa,
                    "m3": m3,
                    "prices": prices
                })
                
    print(f"Parsed {len(parsed_modules)} module rows from workbook.")
    
    # 2. Build the SQL migration script content
    sql_lines = []
    sql_lines.append("-- ==========================================================================")
    sql_lines.append("-- IMPORTAÇÃO DE PREÇOS E MÓDULOS KARAMS 2026 (ABIMAD 42)")
    sql_lines.append(f"-- Gerado automaticamente em conformidade com as regras de integridade do banco")
    sql_lines.append("-- ==========================================================================\n")
    
    sql_lines.append("BEGIN TRANSACTION;\n")
    
    # A. Inactivate all existing modulo_tecido records for Karams
    sql_lines.append("-- 1. Inativar a totalidade dos preços atuais do fornecedor Karams (fornecedor_id = 1)")
    sql_lines.append("UPDATE pi.modulo_tecido ")
    sql_lines.append("SET fl_ativo = false, data_hora_inativacao = NOW() ")
    sql_lines.append("WHERE id_modulo IN (SELECT id FROM pi.modulo WHERE id_fornecedor = 1) AND fl_ativo = true;\n")
    
    # B. Ensure all fabrics exist in DB
    sql_lines.append("-- 2. Garantir que todos os grupos de tecidos existem no banco de dados")
    for f in sorted(list(fabrics_to_ensure)):
        sql_lines.append(f"INSERT INTO pi.tecido (nome) SELECT '{f}' WHERE NOT EXISTS (SELECT 1 FROM pi.tecido WHERE UPPER(nome) = '{f.upper()}');")
    sql_lines.append("")
    
    # C. Ensure all brands exist in DB
    sql_lines.append("-- 3. Garantir que todas as marcas/modelos existem no banco de dados")
    for b in sorted(list(brands_to_ensure)):
        b_esc = b.replace("'", "''")
        sql_lines.append(f"INSERT INTO pi.marca (nome, fl_ativo) SELECT '{b_esc}', true WHERE NOT EXISTS (SELECT 1 FROM pi.marca WHERE UPPER(nome) = '{b.upper()}');")
    sql_lines.append("")
    
    # Track inserted modules in this run to avoid duplicates if any in sheet
    inserted_modules_tracker = set()
    
    sql_lines.append("-- 4. Atualizar módulos existentes ou inserir novos módulos e seus preços")
    
    updates_count = 0
    inserts_count = 0
    prices_count = 0
    
    for m in parsed_modules:
        brand_norm = normalize(m["brand"])
        desc_norm = normalize(m["description"])
        
        # Get brand ID from local cache (assuming it exists, since we check it or ensure it)
        brand_id = brand_map.get(brand_norm)
        
        # Check if matched in database
        matched_db_id = None
        if brand_id:
            for dbm in db_modules:
                if dbm["id_marca"] == brand_id and dbm["id_categoria"] == m["cat_id"] and dbm["descricao_norm"] == desc_norm:
                    # Compare dimensions (width and depth within 2cm)
                    if abs(dbm["largura"] - m["width"]) < 0.02 and abs(dbm["profundidade"] - m["prof"]) < 0.02:
                        matched_db_id = dbm["id"]
                        break
                        
        desc_esc = m["description"].replace("'", "''")
        brand_esc = m["brand"].replace("'", "''")
        
        if matched_db_id:
            # Matched existing module: UPDATE
            sql_lines.append(f"-- Linha {m['row']} planilha | Modulo Existente ID={matched_db_id}")
            sql_lines.append(
                f"UPDATE pi.modulo "
                f"SET largura = {m['width']:.2f}, profundidade = {m['prof']:.2f}, altura = {m['altura']:.2f}, pa = {m['pa']:.2f}, descricao = '{desc_esc}' "
                f"WHERE id = {matched_db_id};"
            )
            updates_count += 1
            
            # Insert prices
            for fab_name, price in m["prices"].items():
                sql_lines.append(
                    f"INSERT INTO pi.modulo_tecido (id_modulo, id_tecido, valor_tecido, fl_ativo, dt_ultima_revisao) "
                    f"VALUES ({matched_db_id}, (SELECT id FROM pi.tecido WHERE UPPER(nome) = '{fab_name.upper()}' LIMIT 1), {price:.3f}, true, NOW());"
                )
                prices_count += 1
        else:
            # New module: INSERT
            module_key = (brand_norm, m["cat_id"], desc_norm, round(m["width"], 2), round(m["prof"], 2))
            
            if module_key not in inserted_modules_tracker:
                sql_lines.append(f"-- Linha {m['row']} planilha | Novo Modulo")
                sql_lines.append(
                    f"INSERT INTO pi.modulo (id_fornecedor, id_categoria, id_marca, descricao, largura, profundidade, altura, pa) "
                    f"VALUES (1, {m['cat_id']}, (SELECT id FROM pi.marca WHERE UPPER(nome) = '{m['brand'].upper()}' LIMIT 1), '{desc_esc}', {m['width']:.2f}, {m['prof']:.2f}, {m['altura']:.2f}, {m['pa']:.2f});"
                )
                inserted_modules_tracker.add(module_key)
                inserts_count += 1
                
            # Insert prices using subquery to look up newly created module ID
            for fab_name, price in m["prices"].items():
                sql_lines.append(
                    f"INSERT INTO pi.modulo_tecido (id_modulo, id_tecido, valor_tecido, fl_ativo, dt_ultima_revisao) "
                    f"VALUES ("
                    f"  (SELECT id FROM pi.modulo WHERE id_fornecedor = 1 AND id_marca = (SELECT id FROM pi.marca WHERE UPPER(nome) = '{m['brand'].upper()}' LIMIT 1) AND descricao = '{desc_esc}' AND largura = {m['width']:.2f} AND profundidade = {m['prof']:.2f} LIMIT 1), "
                    f"  (SELECT id FROM pi.tecido WHERE UPPER(nome) = '{fab_name.upper()}' LIMIT 1), "
                    f"  {price:.3f}, true, NOW()"
                    f");"
                )
                prices_count += 1
                
        sql_lines.append("") # empty line spacing
        
    sql_lines.append("COMMIT;")
    
    # 3. Write SQL to file
    sql_content = "\n".join(sql_lines)
    with open("import_karams_2026.sql", "w", encoding="utf-8") as f:
        f.write(sql_content)
    print(f"\nSQL migration script generated: import_karams_2026.sql")
    print(f"  Updates to existing modules: {updates_count}")
    print(f"  New modules inserted: {inserts_count}")
    print(f"  Prices inserted: {prices_count}")
    
    # 4. Apply the SQL to local database to verify it works
    print("\nApplying SQL migration to local database...")
    try:
        cur.execute(sql_content)
        conn.commit()
        print("Success! SQL script applied successfully to local database without errors.")
    except Exception as e:
        conn.rollback()
        print(f"Error applying SQL to database: {e}")
        raise e
    finally:
        cur.close()
        conn.close()

if __name__ == "__main__":
    main()
