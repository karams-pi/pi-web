import openpyxl

wb = openpyxl.load_workbook('Docs/TESTE1.xlsx', data_only=True)
ws = wb.active

print(f"{'Row':<5} | {'B(2) Marca':<15} | {'C(3) Larg':<10} | {'L(12) Desc':<20} | {'M(13) Tec':<10} | {'N(14) Val':<10}")
print("-" * 80)

rows_to_check = list(range(1, 25))
for row_idx in rows_to_check:
    row = ws[row_idx]
    marca = row[1].value # Col B
    larg = row[2].value # Col C
    desc = row[11].value # Col L
    tecido = row[12].value # Col M
    valor = row[13].value # Col N
    
    marca_s = str(marca)[:15] if marca else ""
    larg_s = str(larg)[:10] if larg else ""
    desc_s = str(desc)[:20] if desc else ""
    tec_s = str(tecido)[:10] if tecido else ""
    val_s = str(valor)[:10] if valor else ""
    
    print(f"{row_idx:<5} | {marca_s:<15} | {larg_s:<10} | {desc_s:<20} | {tec_s:<10} | {val_s:<10}")

# Find DAMMAN again and show its block
print("\n--- DAMMAN BLOCK ---")
for row_idx in range(1, ws.max_row + 1):
    marca = ws.cell(row=row_idx, column=2).value
    if marca and 'DAMMAN' in str(marca).upper():
        for r2 in range(row_idx, row_idx + 10):
            row = ws[r2]
            print(f"{r2:<5} | {str(row[1].value):<15} | {str(row[2].value):<10} | {str(row[11].value):<20} | {str(row[12].value):<10} | {str(row[13].value):<10}")
        break
