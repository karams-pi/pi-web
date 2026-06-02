import openpyxl

def inspect_headers(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    print("A1 Content:")
    print(sheet["A1"].value)
    print("\nJ1 Content:")
    print(sheet["J1"].value)
    
    print("\nRow 45 Content (Total Row?):")
    for col in range(1, 18):
        cell_val = sheet.cell(row=45, column=col).value
        if cell_val is not None:
            print(f"{openpyxl.utils.get_column_letter(col)}45: {cell_val}")
            
    print("\nA47 Content (Bank Details):")
    print(sheet["A47"].value)
    
    print("\nJ47 Content (Product General Data):")
    print(sheet["J47"].value)
    
    print("\nA58 Content:")
    print(sheet["A58"].value)

if __name__ == "__main__":
    inspect_headers(r"c:\Portifólio\pi-web\Docs\New\PROFORMA- SADIMA FRG2026-PO-04.xlsx")
