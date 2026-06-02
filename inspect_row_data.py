import openpyxl

def inspect_rows(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    print("--- Generated Rows (11 to 20) ---")
    for r in range(11, 21):
        row_vals = []
        for c in range(1, 18):
            cell = sheet.cell(row=r, column=c)
            row_vals.append(f"{openpyxl.utils.get_column_letter(c)}: {cell.value}")
        print(f"Row {r}: {', '.join(row_vals)}")

if __name__ == "__main__":
    inspect_rows(r"c:\Portifólio\pi-web\test_export_15.xlsx")
