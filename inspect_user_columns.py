import openpyxl

def inspect_cols(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    print("--- Column Details ---")
    for col in range(1, sheet.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        # Check header cell at row 10
        cell = sheet.cell(row=10, column=col)
        val = cell.value
        fill = cell.fill
        color_hex = None
        if fill and isinstance(fill, openpyxl.styles.PatternFill):
            if fill.start_color and fill.start_color.rgb:
                color_hex = fill.start_color.rgb
        print(f"Col {col_letter} ({col}): Header='{val}' | Fill Color={color_hex}")

if __name__ == "__main__":
    inspect_cols(r"c:\Portifólio\pi-web\Docs\New\PROFORMA- SADIMA FRG2026-PO-04.xlsx")
