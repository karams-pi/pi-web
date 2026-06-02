import openpyxl

def inspect_all_styles(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    with open("styles_dump.txt", "w", encoding="utf-8") as f:
        f.write("--- Cells with Fills ---\n")
        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                fill = cell.fill
                if fill and fill.fill_type and fill.fill_type != 'none':
                    fg = fill.fgColor.rgb if fill.fgColor else None
                    bg = fill.bgColor.rgb if fill.bgColor else None
                    val_str = str(cell.value)[:30] if cell.value is not None else ""
                    f.write(f"Cell {openpyxl.utils.get_column_letter(c)}{r}: val='{val_str}' | type={fill.fill_type} | fg={fg} | bg={bg}\n")

if __name__ == "__main__":
    inspect_all_styles(r"c:\Portifólio\pi-web\Docs\New\PROFORMA- SADIMA FRG2026-PO-04.xlsx")
