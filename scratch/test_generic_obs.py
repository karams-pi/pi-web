import psycopg2
import requests

# 1. Update database
try:
    conn = psycopg2.connect("host=localhost port=5432 dbname=pi_db user=pi password=pi123")
    cur = conn.cursor()
    cur.execute("UPDATE pi.pi_item SET observacao = 'TEST_OBSERVATION_13' WHERE id = 95")
    conn.commit()
    print("Database updated successfully for item 95.")
    cur.close()
    conn.close()
except Exception as e:
    print("DB Update Error:", e)
    exit(1)

# 2. Download export
url = "http://localhost:5000/api/pi/pis/13/excel?currency=EXW&validity=30&lang=PT"
print("Downloading Karams (PI 13)...")
r = requests.get(url)
if r.status_code == 200:
    filename = "test_export_Karams_13_updated.xlsx"
    with open(filename, "wb") as f:
        f.write(r.content)
    print("Saved!")
else:
    print("Download failed:", r.status_code)
    exit(1)

# 3. Read value using openpyxl
import openpyxl
wb = openpyxl.load_workbook("test_export_Karams_13_updated.xlsx")
ws = wb.active
print("Values in Col 13 (OBSERVATION):")
for r in range(1, ws.max_row + 1):
    val = ws.cell(row=r, column=13).value
    if val is not None:
         print(f"Row {r}: {repr(val)}")
