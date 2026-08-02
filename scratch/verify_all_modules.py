import psycopg2
import openpyxl
import re
import unicodedata

def normalize(text):
    if not text:
        return ""
    nfkd_form = unicodedata.normalize('NFKD', text)
    text = "".join([c for c in nfkd_form if not unicodedata.combining(c)])
    text = text.lower().strip()
    text = re.sub(r'\s+', ' ', text)
    text = text.replace('´', "'").replace('’', "'")
    return text

def clean_val(val):
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val == "" or val.lower() == "none" or val == "-" or val == "0":
            return None
        val = val.replace(',', '.')
        try:
            return float(val)
        except ValueError:
            return val
    return float(val)

def clean_desc(desc):
    if desc is None:
        return ""
    desc = str(desc).strip()
    desc = re.sub(r'\s+', ' ', desc)
    return desc

def parse_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    parsed_modules = []
    
    sheet_cat_mapping = {
        "Karam´s Estofados": "Karam´s Estofados",
        "Karam´s Poltronas": "Karam´s Poltronas",
        "Karam´s Puff e Almofadas": "Karam´s Puff e Almofadas",
        "Karam´s Camas": "Karam´s Camas"
    }
    
    for sheet_name, db_cat_name in sheet_cat_mapping.items():
        if sheet_name not in wb.sheetnames:
            continue
        sheet = wb[sheet_name]
        
        current_model = None
        col_mapping = {}
        
        for r in range(1, sheet.max_row + 1):
            b_val = sheet.cell(row=r, column=2).value
            b_clean = clean_desc(b_val)
            
            # Check if it is a header row
            if b_clean and any(x in b_clean.lower() for x in ["descrição", "descriço"]):
                a_val = sheet.cell(row=r, column=1).value
                if a_val:
                    current_model = clean_desc(a_val)
                    
                col_mapping = {}
                for col in range(8, 19):
                    cell_val = sheet.cell(row=r, column=col).value
                    if cell_val:
                        cell_str = str(cell_val).strip().upper()
                        if re.match(r'^G\d+$', cell_str):
                            col_mapping[col] = cell_str
                continue
                
            if not current_model:
                continue
                
            # Check if it is a module row
            c_val = sheet.cell(row=r, column=3).value
            if b_clean != "" and c_val is not None:
                width = clean_val(c_val)
                if not isinstance(width, float):
                    continue
                    
                prof = clean_val(sheet.cell(row=r, column=4).value)
                pa = 0.0
                m3 = 0.0
                altura = 0.0
                
                if sheet_name == "Karam´s Estofados":
                    pa = clean_val(sheet.cell(row=r, column=5).value) or 0.0
                    m3 = clean_val(sheet.cell(row=r, column=6).value) or 0.0
                    altura = clean_val(sheet.cell(row=r, column=7).value) or 0.0
                elif sheet_name in ["Karam´s Poltronas", "Karam´s Puff e Almofadas"]:
                    altura = clean_val(sheet.cell(row=r, column=5).value) or 0.0
                    m3 = clean_val(sheet.cell(row=r, column=6).value) or 0.0
                elif sheet_name == "Karam´s Camas":
                    m3 = clean_val(sheet.cell(row=r, column=5).value) or 0.0
                    altura = clean_val(sheet.cell(row=r, column=6).value) or 0.0
                    
                prices = {}
                for col, fab_name in col_mapping.items():
                    price = clean_val(sheet.cell(row=r, column=col).value)
                    if isinstance(price, float) and price > 0:
                        prices[fab_name] = price
                        
                parsed_modules.append({
                    "sheet": sheet_name,
                    "row": r,
                    "cat_name": db_cat_name,
                    "brand": current_model,
                    "description": b_clean,
                    "width": width,
                    "prof": prof,
                    "altura": altura,
                    "pa": pa,
                    "m3": m3,
                    "prices": prices
                })
                
    return parsed_modules

def run_audit():
    excel_path = r"c:\Portifólio\pi-web\Docs\New\Karams 2026 - Abimad 42 (Exportação).xlsm"
    sheet_modules = parse_excel(excel_path)
    
    conn = psycopg2.connect("host=localhost port=5432 dbname=pi_db user=pi password=pi123")
    cur = conn.cursor()
    
    # Cache DB data to optimize
    cur.execute("SELECT id, nome FROM pi.categoria;")
    cat_map = {normalize(r[1]): r[0] for r in cur.fetchall()}
    
    cur.execute("SELECT id, nome FROM pi.marca;")
    brand_map = {normalize(r[1]): r[0] for r in cur.fetchall()}
    
    cur.execute("SELECT id, nome FROM pi.tecido;")
    tecido_map = {normalize(r[1]): r[0] for r in cur.fetchall()}
    
    # Load all DB modules for Karams
    cur.execute("""
        SELECT m.id, m.id_marca, m.id_categoria, m.descricao, m.largura, m.profundidade, m.altura, m.pa, m.m3
        FROM pi.modulo m
        WHERE m.id_fornecedor = 1;
    """)
    db_modules = []
    for r in cur.fetchall():
        db_modules.append({
            "id": r[0],
            "id_marca": r[1],
            "id_categoria": r[2],
            "descricao_norm": normalize(r[3]),
            "largura": float(r[4]),
            "profundidade": float(r[5]),
            "altura": float(r[6]),
            "pa": float(r[7]),
            "m3": float(r[8])
        })
        
    mismatches = []
    verified_count = 0
    
    for m in sheet_modules:
        brand_norm = normalize(m["brand"])
        cat_norm = normalize(m["cat_name"])
        desc_norm = normalize(m["description"])
        
        brand_id = brand_map.get(brand_norm)
        cat_id = cat_map.get(cat_norm)
        
        if not brand_id:
            mismatches.append(f"Sheet Row {m['row']} ({m['sheet']}): Brand '{m['brand']}' not found in DB.")
            continue
            
        if not cat_id:
            mismatches.append(f"Sheet Row {m['row']} ({m['sheet']}): Category '{m['cat_name']}' not found in DB.")
            continue
            
        # Match module in DB
        matched_mod = None
        for dbm in db_modules:
            if dbm["id_marca"] == brand_id and dbm["id_categoria"] == cat_id and dbm["descricao_norm"] == desc_norm:
                if abs(dbm["largura"] - m["width"]) < 0.02 and abs(dbm["profundidade"] - m["prof"]) < 0.02:
                    matched_mod = dbm
                    break
                    
        if not matched_mod:
            mismatches.append(f"Sheet Row {m['row']} ({m['sheet']}): Module '{m['description']}' (W={m['width']}, D={m['prof']}) for brand '{m['brand']}' not found in DB.")
            continue
            
        # Check module dimensions
        dim_errors = []
        if abs(matched_mod["largura"] - m["width"]) > 0.01:
            dim_errors.append(f"Width mismatch: DB={matched_mod['largura']}, Sheet={m['width']}")
        if abs(matched_mod["profundidade"] - m["prof"]) > 0.01:
            dim_errors.append(f"Depth mismatch: DB={matched_mod['profundidade']}, Sheet={m['prof']}")
        if abs(matched_mod["altura"] - m["altura"]) > 0.01:
            dim_errors.append(f"Height mismatch: DB={matched_mod['altura']}, Sheet={m['altura']}")
        if abs(matched_mod["pa"] - m["pa"]) > 0.01:
            dim_errors.append(f"PA mismatch: DB={matched_mod['pa']}, Sheet={m['pa']}")
            
        if dim_errors:
            mismatches.append(f"Sheet Row {m['row']} ({m['sheet']}): Module ID {matched_mod['id']} dimension mismatch: {', '.join(dim_errors)}")
            
        # Check prices
        for fab_name, price in m["prices"].items():
            fab_norm = normalize(fab_name)
            tecido_id = tecido_map.get(fab_norm)
            if not tecido_id:
                mismatches.append(f"Sheet Row {m['row']} ({m['sheet']}): Fabric '{fab_name}' not found in DB.")
                continue
                
            cur.execute("""
                SELECT id, valor_tecido 
                FROM pi.modulo_tecido 
                WHERE id_modulo = %s AND id_tecido = %s AND fl_ativo = true;
            """, (matched_mod["id"], tecido_id))
            price_rows = cur.fetchall()
            
            if len(price_rows) == 0:
                mismatches.append(f"Sheet Row {m['row']} ({m['sheet']}): Missing active price for Module ID {matched_mod['id']}, Fabric '{fab_name}' (Expected: {price})")
            elif len(price_rows) > 1:
                mismatches.append(f"Sheet Row {m['row']} ({m['sheet']}): Duplicate active prices ({len(price_rows)}) for Module ID {matched_mod['id']}, Fabric '{fab_name}'.")
            else:
                db_price_id, db_price = price_rows[0]
                if abs(float(db_price) - price) > 0.01:
                    mismatches.append(f"Sheet Row {m['row']} ({m['sheet']}): Price mismatch for Module ID {matched_mod['id']}, Fabric '{fab_name}'. DB={db_price}, Sheet={price}")
                    
        verified_count += 1
        
    print(f"\nAudit completed. Checked {verified_count} / {len(sheet_modules)} modules.")
    print(f"Total discrepancies found: {len(mismatches)}")
    
    with open(r"c:\Portifólio\pi-web\scratch\audit_mismatches.txt", "w", encoding="utf-8") as f:
        if mismatches:
            for item in mismatches:
                f.write(item + "\n")
            print("Discrepancies saved to scratch/audit_mismatches.txt")
        else:
            f.write("All modules and prices match 100% correctly!\n")
            print("100% correct matching confirmed!")
            
    cur.close()
    conn.close()

if __name__ == "__main__":
    run_audit()
