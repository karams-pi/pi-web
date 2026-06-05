import openpyxl

def rgb_to_hex(color):
    if color and hasattr(color, 'value') and isinstance(color.value, str):
        val = color.value
        if len(val) == 8:
            return val[2:]
        return val
    return None

def inspect_styles_and_formulas(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    
    sheet = wb[sheet_name]
    sheet_data = wb_data[sheet_name]
    
    print(f"\n==================== SHEET: {sheet_name} (Rows 1 to 20) ====================")
    
    for r in range(1, 21):
        row_str = []
        has_content = False
        for c in range(1, min(sheet.max_column + 1, 20)):
            try:
                cell = sheet.cell(row=r, column=c)
                cell_data = sheet_data.cell(row=r, column=c)
                val = cell_data.value
                formula = cell.value
                
                fill = cell.fill
                fill_color = rgb_to_hex(fill.fgColor) if fill and fill.fill_type == 'solid' else None
                
                font = cell.font
                font_bold = font.bold if font else False
                font_size = font.size if font else None
                font_color = rgb_to_hex(font.color) if font and font.color else None
                
                align = cell.alignment
                halign = align.horizontal if align else None
                
                if val is not None or (formula is not None and str(formula).startswith("=")):
                    has_content = True
                    content = f"'{val}'" if val is not None else ""
                    if formula is not None and str(formula).startswith("="):
                        content += f" [{formula}]"
                    
                    style_str = ""
                    if fill_color:
                        style_str += f" bg={fill_color}"
                    if font_bold:
                        style_str += " B"
                    if font_size and font_size != 11:
                        style_str += f" sz={font_size}"
                    if font_color:
                        style_str += f" fg={font_color}"
                    if halign:
                        style_str += f" align={halign}"
                    
                    row_str.append(f"{openpyxl.utils.get_column_letter(c)}{r}: {content}{style_str}")
            except Exception as e:
                row_str.append(f"{openpyxl.utils.get_column_letter(c)}{r}: err {e}")
        
        if has_content:
            print(f"Row {r}: " + " | ".join(row_str))

if __name__ == "__main__":
    file_path = r"c:\Portifólio\pi-web\Docs\EDC\EDC - RALLI - 11-03-2026.xlsx"
    inspect_styles_and_formulas(file_path, "Resumo 100%")
