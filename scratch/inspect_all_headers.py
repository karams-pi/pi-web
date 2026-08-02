import openpyxl

def main():
    path = r"c:\Portifólio\pi-web\Docs\New\Karams 2026 - Abimad 42 (Exportação).xlsm"
    wb = openpyxl.load_workbook(path, data_only=True)
    
    for sname in wb.sheetnames:
        if "Karam" not in sname:
            continue
        sheet = wb[sname]
        print(f"\nSheet: {sname} (max_row={sheet.max_row}, max_col={sheet.max_column})")
        print("-" * 65)
        for r in range(1, sheet.max_row + 1):
            b_val = sheet.cell(row=r, column=2).value
            if b_val and any(x in str(b_val).lower() for x in ["descrição", "descriço"]):
                a_val = sheet.cell(row=r, column=1).value
                print(f"Row {r:4d} | Col A: {a_val} | Col B: {b_val}")

if __name__ == "__main__":
    main()
