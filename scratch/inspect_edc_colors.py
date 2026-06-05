import openpyxl

def inspect_colors(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"\nSheet: {name}")
        for r in range(1, min(sheet.max_row + 1, 40)):
            for c in range(1, min(sheet.max_column + 1, 30)):
                cell = sheet.cell(row=r, column=c)
                fill = cell.fill
                if fill and fill.fill_type == 'solid' and fill.fgColor and fill.fgColor.rgb:
                    val = str(cell.value)[:20] if cell.value is not None else ""
                    print(f"  Row {r}, Col {openpyxl.utils.get_column_letter(c)} ('{val}'): color={fill.fgColor.rgb}")

if __name__ == "__main__":
    inspect_colors(r"c:\Portifólio\pi-web\Docs\EDC\EDC - RALLI - 11-03-2026.xlsx")
