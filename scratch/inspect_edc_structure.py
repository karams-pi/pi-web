import openpyxl

def inspect_sheets(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    
    print(f"Sheets: {wb.sheetnames}")
    for name in wb.sheetnames:
        sheet = wb[name]
        sheet_data = wb_data[name]
        print(f"\n--- Sheet: {name} ({sheet.max_row} rows x {sheet.max_column} cols) ---")
        
        # Print first few rows to understand structure
        for r in range(1, min(sheet.max_row + 1, 15)):
            row_vals = []
            for c in range(1, min(sheet.max_column + 1, 10)):
                val = sheet_data.cell(row=r, column=c).value
                formula = sheet.cell(row=r, column=c).value
                
                cell_str = ""
                if val is not None:
                    cell_str += f"{val}"
                if formula is not None and str(formula).startswith("="):
                    cell_str += f" ({formula})"
                
                row_vals.append(f"{openpyxl.utils.get_column_letter(c)}{r}: {cell_str}")
            print(" | ".join(row_vals))

if __name__ == "__main__":
    import sys
    inspect_sheets(r"c:\Portifólio\pi-web\Docs\EDC\EDC - RALLI - 11-03-2026.xlsx")
