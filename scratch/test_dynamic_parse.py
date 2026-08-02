import openpyxl
import re

def clean_val(val):
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val == "" or val.lower() == "none" or val == "-" or val == "0":
            return None
        val = val.replace(',', '.')
        try:
            return float(val)
        except ValueError:
            return val
    return float(val)

def clean_desc(desc):
    if desc is None:
        return ""
    desc = str(desc).strip()
    desc = re.sub(r'\s+', ' ', desc)
    return desc

def parse_with_dynamic_headers(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    
    for sname in wb.sheetnames:
        if "Karam" not in sname:
            continue
        sheet = wb[sname]
        print(f"\n=========================================")
        print(f"SHEET: {sname} (max_row={sheet.max_row})")
        print(f"=========================================")
        
        current_model = None
        col_mapping = {}  # col_index -> fabric_name
        
        for r in range(1, sheet.max_row + 1):
            b_val = sheet.cell(row=r, column=2).value
            b_clean = clean_desc(b_val)
            
            # Check if it is a header row
            if b_clean and any(x in b_clean.lower() for x in ["descrição", "descriço"]):
                # This is a header row!
                # Read model name from Column A
                a_val = sheet.cell(row=r, column=1).value
                if a_val:
                    current_model = clean_desc(a_val)
                    
                # Dynamically map columns 8 to 18
                col_mapping = {}
                for col in range(8, 19):
                    cell_val = sheet.cell(row=r, column=col).value
                    if cell_val:
                        cell_str = str(cell_val).strip().upper()
                        if re.match(r'^G\d+$', cell_str):
                            col_mapping[col] = cell_str
                print(f"Row {r:3d}: Model Block '{current_model}' | Fabric Mapping: {col_mapping}")
                continue
                
            if not current_model:
                continue
                
            # Check if it is a module row
            c_val = sheet.cell(row=r, column=3).value
            if b_clean != "" and c_val is not None:
                width = clean_val(c_val)
                if not isinstance(width, float):
                    continue
                    
                # Parse dimensions
                prof = clean_val(sheet.cell(row=r, column=4).value)
                pa = 0.0
                m3 = 0.0
                altura = 0.0
                
                if sname == "Karam´s Estofados":
                    pa = clean_val(sheet.cell(row=r, column=5).value) or 0.0
                    m3 = clean_val(sheet.cell(row=r, column=6).value) or 0.0
                    altura = clean_val(sheet.cell(row=r, column=7).value) or 0.0
                elif sname in ["Karam´s Poltronas", "Karam´s Puff e Almofadas"]:
                    altura = clean_val(sheet.cell(row=r, column=5).value) or 0.0
                    m3 = clean_val(sheet.cell(row=r, column=6).value) or 0.0
                elif sname == "Karam´s Camas":
                    m3 = clean_val(sheet.cell(row=r, column=5).value) or 0.0
                    altura = clean_val(sheet.cell(row=r, column=6).value) or 0.0
                    
                # Read prices using column mapping
                prices = {}
                for col, fab_name in col_mapping.items():
                    price = clean_val(sheet.cell(row=r, column=col).value)
                    if isinstance(price, float) and price > 0:
                        prices[fab_name] = price
                        
                # Only print or count if we have valid dimensions/prices
                print(f"  Row {r:3d} | {b_clean:<40} | W={width:.2f} D={prof:.2f} H={altura:.2f} PA={pa:.2f} M3={m3:.3f} | Prices={prices}")

if __name__ == "__main__":
    parse_with_dynamic_headers(r"c:\Portifólio\pi-web\Docs\New\Karams 2026 - Abimad 42 (Exportação).xlsm")
