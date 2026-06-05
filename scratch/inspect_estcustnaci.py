import openpyxl

def inspect_est_cust_naci(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    
    sheet = wb["Est. Cust. Naci."]
    sheet_data = wb_data["Est. Cust. Naci."]
    
    print("==================== SHEET: Est. Cust. Naci. ====================")
    # Row 2 contains subheaders/group headers
    # Row 3 contains column headers
    # Let's print headers first (A2 to BC3)
    headers = []
    for c in range(1, sheet.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(c)
        group_header = sheet_data.cell(row=2, column=c).value
        col_header = sheet_data.cell(row=3, column=c).value
        
        header_str = ""
        if group_header:
            header_str += f"[{group_header}] "
        header_str += f"{col_header or ''}"
        
        headers.append(f"{col_letter}: {header_str}")
    
    print("Headers:\n" + "\n".join(headers))
    
    # Print row 6 as a data sample with formulas
    print("\nRow 6 (Data sample):")
    row_str = []
    for c in range(1, sheet.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(c)
        val = sheet_data.cell(row=6, column=c).value
        formula = sheet.cell(row=6, column=c).value
        
        cell_val = f"'{val}'" if val is not None else ""
        if formula is not None and str(formula).startswith("="):
            cell_val += f" [{formula}]"
        
        row_str.append(f"{col_letter}: {cell_val}")
    print("\n".join(row_str))

if __name__ == "__main__":
    inspect_est_cust_naci(r"c:\Portifólio\pi-web\Docs\EDC\EDC - RALLI - 11-03-2026.xlsx")
