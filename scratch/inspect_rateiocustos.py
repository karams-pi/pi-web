import openpyxl

def inspect_rateio_custos(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    
    sheet = wb["Rateio Custos Fixos"]
    sheet_data = wb_data["Rateio Custos Fixos"]
    
    print("==================== SHEET: Rateio Custos Fixos ====================")
    # Row 10 contains headers
    # Row 11 contains total values of expenses
    # Row 12 is data
    headers = []
    for c in range(1, sheet.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(c)
        col_header = sheet_data.cell(row=10, column=c).value
        total_val = sheet_data.cell(row=11, column=c).value
        total_formula = sheet.cell(row=11, column=c).value
        
        header_str = f"{col_header or ''}"
        if total_val is not None or total_formula is not None:
            header_str += f" [Total: {total_val}"
            if total_formula is not None and str(total_formula).startswith("="):
                header_str += f" ({total_formula})"
            header_str += "]"
        
        headers.append(f"{col_letter}: {header_str}")
    
    print("Headers:\n" + "\n".join(headers))
    
    print("\nRow 12 (Data sample):")
    row_str = []
    for c in range(1, sheet.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(c)
        val = sheet_data.cell(row=12, column=c).value
        formula = sheet.cell(row=12, column=c).value
        
        cell_val = f"'{val}'" if val is not None else ""
        if formula is not None and str(formula).startswith("="):
            cell_val += f" [{formula}]"
        
        row_str.append(f"{col_letter}: {cell_val}")
    print("\n".join(row_str))

if __name__ == "__main__":
    inspect_rateio_custos(r"c:\Portifólio\pi-web\Docs\EDC\EDC - RALLI - 11-03-2026.xlsx")
