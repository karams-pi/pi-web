import openpyxl

def print_col_2(filename):
    print(f"\n--- Column 2 values in {filename} ---")
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=2).value
        if val is not None:
            print(f"Row {r}: {repr(val)}")

print_col_2("test_export_Ferguile_15.xlsx")
