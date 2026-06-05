import openpyxl

def inspect_lista_de_compras(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    
    sheet = wb["LISTA DE COMPRAS"]
    sheet_data = wb_data["LISTA DE COMPRAS"]
    
    print("==================== SHEET: LISTA DE COMPRAS ====================")
    for r in range(1, 15):
        row_str = []
        has_content = False
        for c in range(1, min(sheet.max_column + 1, 20)):
            cell = sheet.cell(row=r, column=c)
            cell_data = sheet_data.cell(row=r, column=c)
            val = cell_data.value
            formula = cell.value
            
            if val is not None or (formula is not None and str(formula).startswith("=")):
                has_content = True
                content = f"'{val}'"
                if formula is not None and str(formula).startswith("="):
                    content += f" [{formula}]"
                row_str.append(f"{openpyxl.utils.get_column_letter(c)}{r}: {content}")
        if has_content:
            print(f"Row {r}: " + " | ".join(row_str))

if __name__ == "__main__":
    inspect_lista_de_compras(r"c:\Portifólio\pi-web\Docs\EDC\EDC - RALLI - 11-03-2026.xlsx")
