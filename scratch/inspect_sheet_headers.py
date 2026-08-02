import openpyxl

def inspect_sheet_headers(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    for sname in wb.sheetnames:
        if "Karam" not in sname:
            continue
        sheet = wb[sname]
        print(f"\n=========================================")
        print(f"SHEET: {sname}")
        print(f"=========================================")
        # Find first header row (which has 'descrição' or 'descriço' in column B)
        header_row = None
        for r in range(1, sheet.max_row + 1):
            val = sheet.cell(row=r, column=2).value
            if val and any(x in str(val).lower() for x in ["descrição", "descriço"]):
                header_row = r
                break
                
        if header_row:
            print(f"Header found on row {header_row}:")
            headers = []
            for c in range(1, 20):
                val = sheet.cell(row=header_row, column=c).value
                headers.append(f"{openpyxl.utils.get_column_letter(c)}: {val}")
            print(" | ".join(headers))
            
            # Print next 2 rows (data rows)
            print("Data row +1:")
            row_vals = []
            for c in range(1, 20):
                val = sheet.cell(row=header_row + 1, column=c).value
                row_vals.append(f"{openpyxl.utils.get_column_letter(c)}: {val}")
            print(" | ".join(row_vals))
            
            print("Data row +2:")
            row_vals2 = []
            for c in range(1, 20):
                val = sheet.cell(row=header_row + 2, column=c).value
                row_vals2.append(f"{openpyxl.utils.get_column_letter(c)}: {val}")
            print(" | ".join(row_vals2))
        else:
            print("No header found!")

if __name__ == "__main__":
    inspect_sheet_headers(r"c:\Portifólio\pi-web\Docs\New\Karams 2026 - Abimad 42 (Exportação).xlsm")
