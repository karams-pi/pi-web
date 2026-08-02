import openpyxl
import re

def clean_val(val):
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val == "" or val.lower() == "none" or val == "0" or val == "-":
            return 0.0
        # Replace comma with dot
        val = val.replace(',', '.')
        try:
            return float(val)
        except ValueError:
            return val
    return float(val)

def clean_desc(desc):
    if desc is None:
        return ""
    desc = str(desc).strip()
    desc = re.sub(r'\s+', ' ', desc)
    return desc

def parse_sheet(sheet, sname):
    print(f"\nParsing sheet: {sname}")
    current_model = None
    modules = []
    
    # We will iterate row by row
    for r in range(1, sheet.max_row + 1):
        b_val = sheet.cell(row=r, column=2).value
        b_clean = clean_desc(b_val)
        
        # Check if it is a model header row
        if b_clean and any(x in b_clean.lower() for x in ["descrição", "descriço"]):
            a_val = sheet.cell(row=r, column=1).value
            if a_val:
                current_model = clean_desc(a_val)
                print(f"  Line {r:3d}: New Model Block -> '{current_model}'")
            continue
            
        # If we don't have a model block active, skip
        if not current_model:
            continue
            
        # A row is a module row if it has a description in B, and a width in C
        c_val = sheet.cell(row=r, column=3).value
        if b_clean != "" and c_val is not None:
            # Let's see if the width can be parsed as a float
            width = clean_val(c_val)
            if not isinstance(width, float):
                # Probably a header or comment row
                continue
                
            # Parse other columns depending on the sheet
            prof = clean_val(sheet.cell(row=r, column=4).value)
            
            pa = 0.0
            m3 = 0.0
            altura = 0.0
            prices = []
            
            if sname == "Karam´s Estofados":
                # Estofados columns
                pa = clean_val(sheet.cell(row=r, column=5).value)
                m3 = clean_val(sheet.cell(row=r, column=6).value)
                altura = clean_val(sheet.cell(row=r, column=7).value)
                # Prices are in columns 9 to 17 (I to Q)
                for col in range(9, 18):
                    prices.append(clean_val(sheet.cell(row=r, column=col).value))
            elif sname in ["Karam´s Poltronas", "Karam´s Puff e Almofadas"]:
                # Poltronas / Puffs columns
                altura = clean_val(sheet.cell(row=r, column=5).value)
                m3 = clean_val(sheet.cell(row=r, column=6).value)
                pa = 0.0
                # Prices are in columns 8 to 16 (H to P)
                for col in range(8, 17):
                    prices.append(clean_val(sheet.cell(row=r, column=col).value))
            elif sname == "Karam´s Camas":
                # Camas columns
                m3 = clean_val(sheet.cell(row=r, column=5).value)
                altura = clean_val(sheet.cell(row=r, column=6).value)
                pa = 0.0
                # Prices are in columns 8 to 16 (H to P)
                for col in range(8, 17):
                    prices.append(clean_val(sheet.cell(row=r, column=col).value))
                    
            modules.append({
                "row": r,
                "model": current_model,
                "description": b_clean,
                "width": width,
                "prof": prof,
                "pa": pa,
                "m3": m3,
                "altura": altura,
                "prices": prices
            })
            
    print(f"  Parsed {len(modules)} modules in sheet '{sname}'")
    return modules

def main():
    path = r"c:\Portifólio\pi-web\Docs\New\Karams 2026 - Abimad 42 (Exportação).xlsm"
    wb = openpyxl.load_workbook(path, data_only=True)
    
    all_modules = []
    for sname in wb.sheetnames:
        if "Karam" in sname:
            sheet = wb[sname]
            mods = parse_sheet(sheet, sname)
            all_modules.extend(mods)
            
    print(f"\nTotal modules parsed: {len(all_modules)}")
    
    # Save a summary of parsed modules to check
    with open(r"c:\Portifólio\pi-web\scratch\parsed_modules_summary.txt", "w", encoding="utf-8") as f:
        for m in all_modules:
            f.write(f"Row {m['row']:3d} | Model: {m['model']:<25} | Desc: {m['description']:<40} | W={m['width']:.2f} | D={m['prof']:.2f} | H={m['altura']:.2f} | PA={m['pa']:.2f} | M3={m['m3']:.3f} | Prices={m['prices']}\n")

if __name__ == "__main__":
    main()
