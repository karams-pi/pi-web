import openpyxl

def main():
    path = r"c:\Portifólio\pi-web\Docs\New\Karams 2026 - Abimad 42 (Exportação).xlsm"
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb["Karam´s Estofados"]
    
    print("Row | Col A | Col B | Col C | Col I (G0)")
    print("-" * 60)
    for r in range(1, 100):
        a = sheet.cell(row=r, column=1).value
        b = sheet.cell(row=r, column=2).value
        c = sheet.cell(row=r, column=3).value
        i = sheet.cell(row=r, column=9).value
        
        a_str = f"'{a}'" if a is not None else "None"
        b_str = f"'{b}'" if b is not None else "None"
        c_str = f"'{c}'" if c is not None else "None"
        i_str = f"'{i}'" if i is not None else "None"
        
        # print if anything is not None
        if a or b or c or i:
            print(f"{r:3d} | {a_str:<25} | {b_str:<30} | {c_str:<8} | {i_str}")

if __name__ == "__main__":
    main()
