import openpyxl

def print_obs(filename, col_idx):
    print(f"\n--- Observation values in {filename} (Col {col_idx}) ---")
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=col_idx).value
        # If merged cell, it might be the start or None
        if val is not None:
            print(f"Row {r}: {repr(val)}")

print_obs("test_export_Karams_13.xlsx", 13)
print_obs("test_export_Ferguile_15.xlsx", 14)
