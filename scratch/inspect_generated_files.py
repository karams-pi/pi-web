import openpyxl

def inspect_excel(filename):
    print(f"\n--- Inspecting {filename} ---")
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # 1. Print default font size
    # Check sheet view default or first cell font
    print(f"Cell A1 font size: {ws['A1'].font.size if ws['A1'].font else 'None'}")
    print(f"Cell A2 font size: {ws['A2'].font.size if ws['A2'].font else 'None'}")
    print(f"Cell A3 font size: {ws['A3'].font.size if ws['A3'].font else 'None'}")
    print(f"Cell D10 font size: {ws['D10'].font.size if ws['D10'].font else 'None'}")
    
    # 2. Print some row heights
    print("Row Heights:")
    for r in [1, 2, 3, 4, 5, 6, 10, 11, 14, 15, 16]:
        print(f"Row {r}: {ws.row_dimensions[r].height}")
        
    # 3. Print some column widths
    print("Column Widths:")
    for c in range(1, 18):
        col_letter = openpyxl.utils.get_column_letter(c)
        print(f"Col {col_letter}: {ws.column_dimensions[col_letter].width}")

try:
    inspect_excel("test_export_Karams_13.xlsx")
    inspect_excel("test_export_Ferguile_15.xlsx")
except Exception as e:
    print("Error:", e)
