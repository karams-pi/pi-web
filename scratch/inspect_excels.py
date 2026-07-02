import openpyxl
import os
import glob

def inspect_excels():
    files = glob.glob("test_export_*.xlsx")
    print("Found files:", files)
    for f in files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            if "Resumo 100%" in wb.sheetnames:
                sheet = wb["Resumo 100%"]
                print(f"File: {os.path.basename(f)}")
                print(f"  E15 (PRODUTO Unit Price): {sheet['E15'].value}")
                print(f"  D15 (Qty): {sheet['D15'].value}")
                print(f"  F15 (Total FOB USD): {sheet['F15'].value}")
                # Let's calculate total FOB / Qty
                try:
                    qty = float(sheet['D15'].value)
                    total = float(sheet['F15'].value)
                    print(f"  Calculated Unit Price (Total/Qty): {total/qty:.4f}")
                except Exception as e:
                    print("Calc error:", e)
            else:
                print(f"File {f} does not have Resumo 100% sheet. Sheets: {wb.sheetnames}")
            wb.close()
        except Exception as e:
            print("Error reading", f, ":", e)

if __name__ == "__main__":
    inspect_excels()
