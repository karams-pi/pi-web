import openpyxl

def dump_sheet(sheet, name, f):
    f.write(f"\n=========================================\n")
    f.write(f"SHEET: {name}\n")
    f.write(f"=========================================\n")
    
    # Let's inspect merged cells in this sheet
    f.write("Merged ranges:\n")
    for r in sheet.merged_cells.ranges:
        f.write(f"  {r}\n")
    f.write("\n")
    
    # Write first 50 rows
    for r in range(1, 55):
        row_vals = []
        for c in range(1, 20):  # columns A to S
            cell = sheet.cell(row=r, column=c)
            val = cell.value
            if val is not None:
                row_vals.append(f"{openpyxl.utils.get_column_letter(c)}{r}: '{val}'")
            else:
                row_vals.append(f"{openpyxl.utils.get_column_letter(c)}{r}: None")
        f.write(f"Row {r:02d}: {', '.join(row_vals)}\n")

def main():
    path = r"c:\Portifólio\pi-web\Docs\New\Karams 2026 - Abimad 42 (Exportação).xlsm"
    wb = openpyxl.load_workbook(path, data_only=True)
    
    with open(r"c:\Portifólio\pi-web\scratch\excel_sheets_dump.txt", "w", encoding="utf-8") as f:
        for sheetname in wb.sheetnames:
            if "Karam" in sheetname:
                dump_sheet(wb[sheetname], sheetname, f)
                
    print("Dump completed. View c:\\Portifólio\\pi-web\\scratch\\excel_sheets_dump.txt")

if __name__ == "__main__":
    main()
