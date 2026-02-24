import openpyxl

wb = openpyxl.load_workbook('Docs/TESTE1.xlsx', data_only=True)
ws = wb.active

print(f"Columns: A=1, B=2 (MARCA), ..., L=12 (DESC), M=13 (TECIDO), N=14 (VALOR)")

# Inspect DAMMAN rows
target = 'DAMMAN'
seen_rows = 0
for row_idx in range(1, ws.max_row + 1):
    marca = str(ws.cell(row=row_idx, column=2).value or '').strip()
    desc = str(ws.cell(row=row_idx, column=12).value or '').strip()
    tecido = str(ws.cell(row=row_idx, column=13).value or '').strip()
    valor = ws.cell(row=row_idx, column=14).value
    
    if target in marca.upper() or target in desc.upper():
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 16)]
        print(f"Row {row_idx}: {row_vals}")
        seen_rows += 1
        if seen_rows > 20: break

print("--- End of DAMMAN inspection ---")
